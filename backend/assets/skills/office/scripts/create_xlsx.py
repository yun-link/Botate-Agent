#!/usr/bin/env python3
"""
create_xlsx.py - 从JSON创建Excel文件，与read_xlsx.py格式完全兼容
Usage: python create_xlsx.py input.json [output.xlsx]
"""

import sys
import json
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import TYPE_FORMULA
from copy import copy


def parse_color(color_str):
    """解析颜色字符串为RGB元组"""
    if not color_str or color_str == 'None':
        return None
    
    try:
        # 处理带Alpha通道的ARGB格式（如 FFFF0000）或纯RGB（FF0000）
        color_str = str(color_str).strip()
        if len(color_str) >= 6:
            # 取最后6位作为RGB
            rgb = color_str[-6:]
            return (
                int(rgb[0:2], 16),
                int(rgb[2:4], 16),
                int(rgb[4:6], 16)
            )
    except (ValueError, TypeError):
        pass
    
    return None


def apply_font(cell, font_info):
    """应用字体样式"""
    if not font_info:
        return
    
    font = cell.font
    
    # 创建新Font对象（openpyxl样式不可变，需要重建）
    new_font = Font(
        name=font_info.get('name'),
        size=font_info.get('size'),
        bold=font_info.get('bold'),
        italic=font_info.get('italic'),
        underline=font_info.get('underline') if font_info.get('underline') else None,
        strike=font_info.get('strike'),
        color=parse_color(font_info.get('color')),
        vertAlign='superscript' if font_info.get('superscript') else 
                  'subscript' if font_info.get('subscript') else None
    )
    
    cell.font = new_font


def apply_fill(cell, fill_info):
    """应用填充/背景色"""
    if not fill_info:
        return
    
    fill_type = fill_info.get('type')
    
    if fill_type == 'solid':
        color = parse_color(fill_info.get('color'))
        if color:
            # openpyxl使用aRGB，需要添加Alpha通道
            fill = PatternFill(
                start_color=f'FF{fill_info.get("color", "FFFFFF")[-6:]}',
                end_color=f'FF{fill_info.get("color", "FFFFFF")[-6:]}',
                fill_type='solid'
            )
            cell.fill = fill
    elif fill_type == 'gradientFill':
        # 简化处理，渐变填充转为纯色或默认
        pass


def apply_border(cell, border_info):
    """应用边框"""
    if not border_info:
        return
    
    def create_side(side_info):
        if not side_info:
            return None
        
        style = side_info.get('style')
        if not style or style == 'None':
            return None
        
        color = parse_color(side_info.get('color'))
        return Side(
            style=style,
            color=f'FF{side_info.get("color", "000000")[-6:]}' if color else 'FF000000'
        )
    
    new_border = Border(
        left=create_side(border_info.get('left')),
        right=create_side(border_info.get('right')),
        top=create_side(border_info.get('top')),
        bottom=create_side(border_info.get('bottom')),
        diagonal=create_side(border_info.get('diagonal'))
    )
    
    cell.border = new_border


def apply_alignment(cell, alignment_info):
    """应用对齐方式"""
    if not alignment_info:
        return
    
    new_alignment = Alignment(
        horizontal=alignment_info.get('horizontal'),
        vertical=alignment_info.get('vertical'),
        wrap_text=alignment_info.get('wrap_text'),
        shrink_to_fit=alignment_info.get('shrink_to_fit'),
        indent=alignment_info.get('indent'),
        text_rotation=alignment_info.get('text_rotation')
    )
    
    cell.alignment = new_alignment


def apply_protection(cell, protection_info):
    """应用保护设置"""
    if not protection_info:
        return
    
    new_protection = Protection(
        locked=protection_info.get('locked', True),
        hidden=protection_info.get('hidden', False)
    )
    
    cell.protection = new_protection


def apply_style(cell, style_info):
    """应用完整样式"""
    if not style_info:
        return
    
    apply_font(cell, style_info.get('font'))
    apply_fill(cell, style_info.get('fill'))
    apply_border(cell, style_info.get('border'))
    apply_alignment(cell, style_info.get('alignment'))
    apply_protection(cell, style_info.get('protection'))


def set_value_and_type(cell, value, cell_type, number_format=None):
    """设置单元格值和类型"""
    # 处理公式
    if cell_type == 'formula' and isinstance(value, str) and value.startswith('='):
        cell.value = value
    else:
        cell.value = value
    
    # 应用数字格式
    if number_format:
        cell.number_format = number_format


def create_sheet(wb, sheet_data, expand_merged=False):
    """
    从JSON数据创建工作表
    
    expand_merged: 如果JSON是展开模式（ghost cells也有完整数据），
                   需要特殊处理避免重复写入
    """
    sheet_name = sheet_data.get('name', 'Sheet1')
    
    # 创建工作表
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
    
    # 清空默认内容
    ws.delete_rows(1, ws.max_row)
    
    rows_data = sheet_data.get('rows', [])
    if not rows_data:
        return ws
    
    # 第一步：写入所有普通单元格和主单元格的数据
    main_cells = {}  # 记录主单元格位置 (row, col) -> value
    
    for row_data in rows_data:
        row_num = row_data.get('row_number', 1)
        
        for cell_data in row_data.get('cells', []):
            if not cell_data:
                continue
            
            # 跳过幽灵单元格（它们不独立存在）
            if cell_data.get('ghost_cell') and not expand_merged:
                continue
            
            row = cell_data.get('row', row_num)
            col = cell_data.get('column', 1)
            coordinate = cell_data.get('coordinate')
            
            # 获取值和类型
            value = cell_data.get('value')
            cell_type = cell_data.get('type', 'string')
            number_format = cell_data.get('number_format')
            
            # 定位单元格
            cell = ws.cell(row=row, column=col)
            
            # 设置值
            set_value_and_type(cell, value, cell_type, number_format)
            
            # 应用样式（仅主单元格或普通单元格）
            if not cell_data.get('ghost_cell'):
                apply_style(cell, cell_data.get('style'))
            
            # 记录主单元格信息用于后续合并
            if cell_data.get('is_merged') and cell_data.get('merged_range'):
                merged_range = cell_data['merged_range']
                if merged_range.get('is_main_cell'):
                    main_cells[(row, col)] = {
                        'range': merged_range,
                        'end_row': merged_range['end_row'],
                        'end_col': merged_range['end_col']
                    }
    
    # 第二步：处理合并单元格
    # 收集所有需要合并的区域
    merge_ranges = sheet_data.get('merged_ranges', [])
    
    for merge_info in merge_ranges:
        range_str = merge_info.get('range')  # 如 "B2:E4"
        if range_str:
            try:
                ws.merge_cells(range_str)
            except Exception as e:
                print(f"  警告: 合并单元格失败 {range_str}: {e}")
    
    # 第三步：应用行列维度
    column_dims = sheet_data.get('column_dimensions', {})
    for col_letter, dim_info in column_dims.items():
        if dim_info.get('width'):
            ws.column_dimensions[col_letter].width = dim_info['width']
        if dim_info.get('hidden'):
            ws.column_dimensions[col_letter].hidden = True
    
    row_dims = sheet_data.get('row_dimensions', {})
    for row_num, dim_info in row_dims.items():
        row_idx = int(row_num) if isinstance(row_num, str) else row_num
        if dim_info.get('height'):
            ws.row_dimensions[row_idx].height = dim_info['height']
        if dim_info.get('hidden'):
            ws.row_dimensions[row_idx].hidden = True
    
    # 第四步：冻结窗格
    freeze_panes = sheet_data.get('freeze_panes')
    if freeze_panes:
        row = freeze_panes.get('row')
        col = freeze_panes.get('col')
        if row or col:
            # openpyxl冻结的是滚动起始单元格
            freeze_row = row + 1 if row else 1
            freeze_col = col + 1 if col else 1
            ws.freeze_panes = ws.cell(row=freeze_row, column=freeze_col).coordinate
    
    # 第五步：工作表属性
    sheet_props = sheet_data.get('sheet_properties', {})
    if sheet_props.get('tab_color'):
        color = parse_color(sheet_props['tab_color'])
        if color:
            ws.sheet_properties.tabColor.rgb = sheet_props['tab_color']
    
    # 自动筛选
    if sheet_props.get('auto_filter'):
        ws.auto_filter.ref = sheet_props['auto_filter']
    
    return ws


def create_xlsx_from_json(json_path, output_path=None):
    """从JSON创建Excel文件"""
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # 创建工作簿
    wb = Workbook()
    
    # 删除默认的Sheet（如果JSON中定义了工作表）
    if data.get('sheets') and len(data['sheets']) > 0:
        wb.remove(wb.active)
    
    # 创建所有工作表
    sheets_data = data.get('sheets', [])
    for i, sheet_data in enumerate(sheets_data):
        print(f"  创建工作表: {sheet_data.get('name', f'Sheet{i+1}')}")
        create_sheet(wb, sheet_data)
    
    # 设置活动工作表
    active_sheet_name = data.get('active_sheet')
    if active_sheet_name and active_sheet_name in wb.sheetnames:
        wb.active = wb[active_sheet_name]
    
    # 保存
    output = output_path or json_path.replace('.json', '.xlsx')
    wb.save(output)
    wb.close()
    
    return output


def verify_round_trip(original_xlsx, json_path, regenerated_xlsx):
    """
    验证双向一致性
    对比：工作表数量、行列数、合并单元格数、关键单元格值
    """
    from openpyxl import load_workbook as load_wb
    
    orig = load_wb(original_xlsx, data_only=True)
    regen = load_wb(regenerated_xlsx, data_only=True)
    
    issues = []
    passed = True
    
    # 1. 工作表数量
    if len(orig.sheetnames) != len(regen.sheetnames):
        issues.append(f"工作表数量: 原{len(orig.sheetnames)} vs 新{len(regen.sheetnames)}")
        passed = False
    
    # 2. 逐个对比工作表
    for sheet_name in orig.sheetnames:
        if sheet_name not in regen.sheetnames:
            issues.append(f"缺失工作表: {sheet_name}")
            passed = False
            continue
        
        orig_ws = orig[sheet_name]
        regen_ws = regen[sheet_name]
        
        # 对比维度
        if orig_ws.max_row != regen_ws.max_row or orig_ws.max_column != regen_ws.max_column:
            issues.append(f"[{sheet_name}] 尺寸: 原({orig_ws.max_row},{orig_ws.max_column}) vs 新({regen_ws.max_row},{regen_ws.max_column})")
            passed = False
        
        # 对比合并单元格数量
        orig_merged = len(list(orig_ws.merged_cells.ranges))
        regen_merged = len(list(regen_ws.merged_cells.ranges))
        if orig_merged != regen_merged:
            issues.append(f"[{sheet_name}] 合并单元格: 原{orig_merged} vs 新{regen_merged}")
            passed = False
        
        # 采样对比关键单元格（ corners, merged main cells）
        sample_cells = [
            (1, 1),  # A1
            (orig_ws.max_row, 1),  # 左下角
            (1, orig_ws.max_column),  # 右上角
            (orig_ws.max_row, orig_ws.max_column),  # 右下角
        ]
        
        # 添加合并区域的主单元格
        for merged_range in orig_ws.merged_cells.ranges:
            sample_cells.append((merged_range.min_row, merged_range.min_col))
        
        for r, c in sample_cells:
            try:
                orig_val = orig_ws.cell(r, c).value
                regen_val = regen_ws.cell(r, c).value
                
                # 处理浮点数精度
                if isinstance(orig_val, float) and isinstance(regen_val, float):
                    match = abs(orig_val - regen_val) < 0.0001
                else:
                    match = orig_val == regen_val
                
                if not match:
                    issues.append(f"[{sheet_name}] 单元格({r},{c}): 原{orig_val} vs 新{regen_val}")
                    passed = False
            except Exception as e:
                issues.append(f"[{sheet_name}] 单元格({r},{c})读取错误: {e}")
                passed = False
    
    orig.close()
    regen.close()
    
    return passed, issues


def main():
    parser = argparse.ArgumentParser(
        description='从JSON创建Excel文件，与read_xlsx.py格式兼容',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python create_xlsx.py data.json
  python create_xlsx.py data.json output.xlsx
        """
    )
    parser.add_argument('input', help='输入JSON文件路径')
    parser.add_argument('output', nargs='?', help='输出XLSX文件路径（可选）')
    
    args = parser.parse_args()
    
    try:
        print(f"正在创建: {args.input}")
        result_path = create_xlsx_from_json(args.input, args.output)
        print(f"✓ 创建成功: {result_path}")
        
        # 显示统计
        from openpyxl import load_workbook
        wb = load_workbook(result_path, data_only=True)
        print(f"  工作表: {len(wb.sheetnames)} 个")
        for name in wb.sheetnames:
            ws = wb[name]
            merged_count = len(list(ws.merged_cells.ranges))
            print(f"    - {name}: {ws.max_row}行 x {ws.max_column}列, {merged_count}个合并区域")
        wb.close()
        
        
    except FileNotFoundError:
        print(f"✗ 错误: 文件不存在 {args.input}")
        sys.exit(1)
    except json.JSONDecodeError as e:
        print(f"✗ 错误: JSON格式无效: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"✗ 错误: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()