#!/usr/bin/env python3
"""
xlsx2json.py - 将Excel文件转换为JSON格式，智能处理合并单元格
Usage: python xlsx2json.py input.xlsx [output.json] [--sheets=all] [--expand-merged]
"""

import sys
import json
import argparse
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Alignment, Protection
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import TYPE_FORMULA, TYPE_NUMERIC, TYPE_STRING, TYPE_BOOL, TYPE_ERROR


def get_cell_type(cell):
    """获取单元格数据类型"""
    if cell.data_type == TYPE_FORMULA:
        return 'formula'
    elif cell.data_type == TYPE_NUMERIC:
        return 'number'
    elif cell.data_type == TYPE_STRING:
        return 'string'
    elif cell.data_type == TYPE_BOOL:
        return 'boolean'
    elif cell.data_type == TYPE_ERROR:
        return 'error'
    return 'unknown'


def get_font_info(font):
    """提取字体信息"""
    if not font:
        return None
    
    color = None
    if font.color and font.color.type == 'rgb':
        color = font.color.rgb
    
    return {
        'name': font.name,
        'size': font.size,
        'bold': font.bold,
        'italic': font.italic,
        'underline': font.underline if font.underline else None,
        'strike': font.strike,
        'color': color,
        'superscript': font.vertAlign == 'superscript',
        'subscript': font.vertAlign == 'subscript'
    }


def get_fill_info(fill):
    """提取填充/背景色信息"""
    if not fill or fill.fill_type is None:
        return None
    
    if fill.fill_type == 'solid':
        return {
            'type': 'solid',
            'color': fill.fgColor.rgb if fill.fgColor else None
        }
    elif fill.fill_type == 'gradientFill':
        return {'type': 'gradient'}
    
    return {'type': str(fill.fill_type)}


def get_border_info(border):
    """提取边框信息"""
    if not border:
        return None
    
    def get_side(side):
        if not side or side.style is None:
            return None
        return {
            'style': side.style,
            'color': side.color.rgb if side.color else None
        }
    
    return {
        'left': get_side(border.left),
        'right': get_side(border.right),
        'top': get_side(border.top),
        'bottom': get_side(border.bottom),
        'diagonal': get_side(border.diagonal) if border.diagonal else None
    }


def get_alignment_info(alignment):
    """提取对齐信息"""
    if not alignment:
        return None
    
    return {
        'horizontal': alignment.horizontal,
        'vertical': alignment.vertical,
        'wrap_text': alignment.wrapText,
        'shrink_to_fit': alignment.shrinkToFit,
        'indent': alignment.indent,
        'text_rotation': alignment.textRotation
    }


def get_number_format(cell):
    """获取数字格式（保留日期、货币等格式）"""
    if cell.number_format and cell.number_format != 'General':
        return cell.number_format
    return None


def parse_cell(cell, merged_ranges_map):
    """
    解析单元格，智能处理合并单元格
    
    merged_ranges_map: 字典，key为合并区域坐标，value为合并区域信息
    """
    coordinate = cell.coordinate
    
    # 检查是否是合并单元格的一部分
    merged_info = merged_ranges_map.get(coordinate)
    
    cell_data = {
        'coordinate': coordinate,
        'row': cell.row,
        'column': cell.column,
        'column_letter': get_column_letter(cell.column),
        'value': cell.value,
        'type': get_cell_type(cell),
        'number_format': get_number_format(cell),
        'is_merged': merged_info is not None if merged_info else False
    }
    
    # 如果是合并单元格，添加合并信息
    if merged_info:
        cell_data['merged_range'] = {
            'start_row': merged_info['min_row'],
            'start_col': merged_info['min_col'],
            'end_row': merged_info['max_row'],
            'end_col': merged_info['max_col'],
            'width': merged_info['max_col'] - merged_info['min_col'] + 1,
            'height': merged_info['max_row'] - merged_info['min_row'] + 1,
            'is_main_cell': merged_info['main_cell'] == coordinate
        }
        
        # 如果不是主单元格，标记为"幽灵单元格"（视觉上属于合并区域但不独立存在）
        if merged_info['main_cell'] != coordinate:
            cell_data['ghost_cell'] = True
            cell_data['value'] = None  # 幽灵单元格的值实际存储在主单元格
    
    # 提取样式信息（仅主单元格或普通单元格）
    if not cell_data.get('ghost_cell'):
        cell_data['style'] = {
            'font': get_font_info(cell.font),
            'fill': get_fill_info(cell.fill),
            'border': get_border_info(cell.border),
            'alignment': get_alignment_info(cell.alignment),
            'protection': {
                'locked': cell.protection.locked,
                'hidden': cell.protection.hidden
            } if cell.protection else None
        }
        
        # 计算值（公式计算结果）
        if cell.data_type == TYPE_FORMULA:
            cell_data['calculated_value'] = cell.value
    
    return cell_data


def build_merged_ranges_map(worksheet):
    """
    构建合并单元格映射表，用于快速查询任意单元格是否属于合并区域
    """
    merged_map = {}
    
    for merged_range in worksheet.merged_cells.ranges:
        min_row, min_col, max_row, max_col = (
            merged_range.min_row,
            merged_range.min_col,
            merged_range.max_row,
            merged_range.max_col
        )
        
        # 主单元格（左上角）坐标
        main_cell = f"{get_column_letter(min_col)}{min_row}"
        
        range_info = {
            'min_row': min_row,
            'min_col': min_col,
            'max_row': max_row,
            'max_col': max_col,
            'main_cell': main_cell
        }
        
        # 将区域内所有单元格坐标映射到该合并区域信息
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                coord = f"{get_column_letter(col)}{row}"
                merged_map[coord] = range_info
    
    return merged_map


def parse_sheet(worksheet, expand_merged=False):
    """
    解析工作表
    
    expand_merged: 如果为True，将合并单元格展开为矩阵形式（每个物理单元格都包含完整数据）
                  如果为False，仅主单元格包含数据，其他标记为ghost_cell
    """
    merged_map = build_merged_ranges_map(worksheet)
    
    # 获取使用范围
    if worksheet.dimensions:
        min_row, max_row, min_col, max_col = (
            worksheet.min_row, worksheet.max_row,
            worksheet.min_col, worksheet.max_col
        )
    else:
        return {
            'name': worksheet.title,
            'dimensions': None,
            'rows': [],
            'merged_ranges': [],
            'max_row': 0,
            'max_col': 0
        }
    
    # 解析所有行
    rows = []
    for row_idx in range(min_row, max_row + 1):
        row_data = []
        for col_idx in range(min_col, max_col + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell_data = parse_cell(cell, merged_map)
            
            # 如果选择展开模式，且是幽灵单元格，复制主单元格的数据
            if expand_merged and cell_data.get('ghost_cell'):
                main_coord = cell_data['merged_range']['main_cell']
                main_cell = worksheet[main_coord]
                # 重新解析主单元格，但保留当前坐标
                main_data = parse_cell(main_cell, merged_map)
                main_data['coordinate'] = cell_data['coordinate']
                main_data['row'] = cell_data['row']
                main_data['column'] = cell_data['column']
                main_data['column_letter'] = cell_data['column_letter']
                main_data['ghost_cell'] = True  # 仍然标记为幽灵
                main_data['merged_range'] = cell_data['merged_range']
                cell_data = main_data
            
            row_data.append(cell_data)
        rows.append({
            'row_number': row_idx,
            'cells': row_data
        })
    
    # 提取合并区域列表（便于快速查看）
    merged_ranges = []
    seen_ranges = set()
    for range_info in merged_map.values():
        key = (range_info['min_row'], range_info['min_col'], 
               range_info['max_row'], range_info['max_col'])
        if key not in seen_ranges:
            seen_ranges.add(key)
            merged_ranges.append({
                'range': f"{range_info['main_cell']}:{get_column_letter(range_info['max_col'])}{range_info['max_row']}",
                'start': range_info['main_cell'],
                'end': f"{get_column_letter(range_info['max_col'])}{range_info['max_row']}",
                'width': range_info['max_col'] - range_info['min_col'] + 1,
                'height': range_info['max_row'] - range_info['min_row'] + 1,
                'value': worksheet[range_info['main_cell']].value
            })
    
    # 提取列宽信息
    column_dimensions = {}
    for col_letter, dim in worksheet.column_dimensions.items():
        if dim.width:
            column_dimensions[col_letter] = {
                'width': dim.width,
                'auto_size': dim.auto_size,
                'hidden': dim.hidden
            }
    
    # 提取行高信息
    row_dimensions = {}
    for row_num, dim in worksheet.row_dimensions.items():
        if dim.height:
            row_dimensions[row_num] = {
                'height': dim.height,
                'hidden': dim.hidden
            }
    
    return {
        'name': worksheet.title,
        'dimensions': {
            'min_row': min_row,
            'max_row': max_row,
            'min_col': min_col,
            'max_col': max_col,
            'used_range': worksheet.dimensions
        },
        'rows': rows,
        'merged_ranges': merged_ranges,
        'merged_count': len(merged_ranges),
        'column_dimensions': column_dimensions,
        'row_dimensions': row_dimensions,
        'freeze_panes': {
            'row': worksheet.freeze_panes[0] if worksheet.freeze_panes else None,
            'col': worksheet.freeze_panes[1] if worksheet.freeze_panes else None
        } if worksheet.freeze_panes else None,
        'sheet_properties': {
            'tab_color': worksheet.sheet_properties.tabColor.rgb if worksheet.sheet_properties.tabColor else None,
            'auto_filter': worksheet.auto_filter.ref if worksheet.auto_filter else None
        }
    }


def parse_xlsx(file_path, sheet_names=None, expand_merged=False):
    """解析Excel文件"""
    wb = load_workbook(file_path, data_only=True)
    
    # 确定要处理的工作表
    if sheet_names == ['all'] or sheet_names is None:
        sheets_to_parse = wb.sheetnames
    else:
        sheets_to_parse = [s for s in sheet_names if s in wb.sheetnames]
        missing = set(sheet_names) - set(wb.sheetnames)
        if missing:
            print(f"警告: 工作表不存在: {', '.join(missing)}")
    
    result = {
        'file': file_path,
        'sheet_names': wb.sheetnames,
        'active_sheet': wb.active.title,
        'sheets': []
    }
    
    for sheet_name in sheets_to_parse:
        ws = wb[sheet_name]
        sheet_data = parse_sheet(ws, expand_merged)
        result['sheets'].append(sheet_data)
        print(f"  ✓ 解析工作表: {sheet_name} ({len(sheet_data['rows'])} 行, {len(sheet_data['merged_ranges'])} 个合并区域)")
    
    wb.close()
    return result


def generate_matrix_view(sheet_data):
    """
    生成矩阵视图（二维数组），便于快速查看数据
    智能处理合并单元格：合并区域只显示一次值
    """
    if not sheet_data['rows']:
        return []
    
    matrix = []
    merged_tracker = {}  # 跟踪合并区域填充状态
    
    for row in sheet_data['rows']:
        matrix_row = []
        for cell in row['cells']:
            if cell.get('ghost_cell'):
                # 如果是幽灵单元格，检查是否已显示
                merge_key = (
                    cell['merged_range']['start_row'],
                    cell['merged_range']['start_col']
                )
                if merge_key in merged_tracker:
                    matrix_row.append(None)  # 已显示过，置空
                else:
                    # 首次遇到该合并区域，显示值
                    matrix_row.append(cell['value'])
                    merged_tracker[merge_key] = True
            else:
                matrix_row.append(cell['value'])
        matrix.append(matrix_row)
    
    return matrix


def main():
    parser = argparse.ArgumentParser(
        description='将Excel文件转换为JSON，智能保留格式和合并单元格',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python xlsx2json.py data.xlsx
  python xlsx2json.py data.xlsx output.json --sheets Sheet1,Sheet2
  python xlsx2json.py data.xlsx output.json --expand-merged
        """
    )
    parser.add_argument('input', help='输入Excel文件路径')
    parser.add_argument('output', nargs='?', help='输出JSON文件路径（可选）')
    parser.add_argument('--sheets', default='all', 
                       help='要转换的工作表，逗号分隔（默认all）')
    parser.add_argument('--expand-merged', action='store_true',
                       help='展开合并单元格（每个物理单元格都包含数据）')
    parser.add_argument('--matrix', action='store_true',
                       help='额外生成矩阵视图（纯数据二维数组）')
    
    args = parser.parse_args()
    
    output_path = args.output or args.input.replace('.xlsx', '.json')
    sheet_list = args.sheets.split(',') if args.sheets != 'all' else ['all']
    
    try:
        print(f"正在解析: {args.input}")
        result = parse_xlsx(args.input, sheet_list, args.expand_merged)
        
        # 可选：生成矩阵视图
        if args.matrix:
            for sheet in result['sheets']:
                sheet['matrix_view'] = generate_matrix_view(sheet)
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2, default=str)
        
        print(f"\n✓ 转换成功: {output_path}")
        print(f"  工作表总数: {len(result['sheet_names'])}")
        print(f"  已转换: {len(result['sheets'])} 个")
        
        # 统计信息
        total_merged = sum(s['merged_count'] for s in result['sheets'])
        if total_merged > 0:
            print(f"  合并单元格: {total_merged} 个区域")
        
    except FileNotFoundError:
        print(f"✗ 错误: 文件不存在 {args.input}")
        sys.exit(1)
    except Exception as e:
        print(f"✗ 错误: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()